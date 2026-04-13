"""``read_spreadsheet`` tool: XLSX, XLS, ODS, TSV, CSV."""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from agents.general.file_tools import resolve_attachment

logger = logging.getLogger("FileTools.read_spreadsheet")


def _read_xlsx(path: Path, sheet_name: Optional[str], max_rows: int) -> Dict[str, Any]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    selected = sheet_name or sheet_names[0]
    if selected not in sheet_names:
        return {"error": {"code": "parse_failed", "message": f"sheet {selected!r} not found"}}
    ws = wb[selected]

    rows: List[List[Any]] = []
    columns: List[str] = []
    truncated = False
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            columns = ["" if c is None else str(c) for c in row]
            continue
        if len(rows) >= max_rows:
            truncated = True
            break
        rows.append(list(row))
    row_count = ws.max_row - 1 if ws.max_row else len(rows)
    return {
        "sheet_name": selected,
        "sheet_names": sheet_names,
        "columns": columns,
        "rows": rows,
        "row_count": row_count,
        "truncated": truncated,
    }


def _read_xls(path: Path, sheet_name: Optional[str], max_rows: int) -> Dict[str, Any]:
    import xlrd  # type: ignore

    book = xlrd.open_workbook(str(path))
    sheet_names = book.sheet_names()
    selected = sheet_name or sheet_names[0]
    if selected not in sheet_names:
        return {"error": {"code": "parse_failed", "message": f"sheet {selected!r} not found"}}
    sh = book.sheet_by_name(selected)
    columns = [str(sh.cell_value(0, c)) for c in range(sh.ncols)] if sh.nrows else []
    rows = []
    truncated = False
    for r in range(1, sh.nrows):
        if len(rows) >= max_rows:
            truncated = True
            break
        rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return {
        "sheet_name": selected,
        "sheet_names": sheet_names,
        "columns": columns,
        "rows": rows,
        "row_count": max(0, sh.nrows - 1),
        "truncated": truncated,
    }


def _read_ods(path: Path, sheet_name: Optional[str], max_rows: int) -> Dict[str, Any]:
    from odf.opendocument import load  # type: ignore
    from odf.table import Table, TableRow, TableCell  # type: ignore
    from odf import text as odftext  # type: ignore

    doc = load(str(path))
    tables = doc.getElementsByType(Table)
    sheet_names = [t.getAttribute("name") for t in tables]
    if not tables:
        return {"sheet_name": None, "sheet_names": [], "columns": [], "rows": [],
                "row_count": 0, "truncated": False}
    selected_name = sheet_name or sheet_names[0]
    table = next((t for t in tables if t.getAttribute("name") == selected_name), tables[0])

    def _row_to_cells(row) -> List[Any]:
        cells: List[Any] = []
        for cell in row.getElementsByType(TableCell):
            ps = cell.getElementsByType(odftext.P)
            text = "".join(p.firstChild.data if p.firstChild else "" for p in ps)
            cells.append(text)
        return cells

    all_rows = [_row_to_cells(r) for r in table.getElementsByType(TableRow)]
    columns = all_rows[0] if all_rows else []
    body = all_rows[1:] if len(all_rows) > 1 else []
    truncated = len(body) > max_rows
    if truncated:
        body = body[:max_rows]
    return {
        "sheet_name": selected_name,
        "sheet_names": sheet_names,
        "columns": columns,
        "rows": body,
        "row_count": len(all_rows) - 1 if all_rows else 0,
        "truncated": truncated,
    }


def _read_delimited(path: Path, delimiter: str, max_rows: int) -> Dict[str, Any]:
    columns: List[str] = []
    rows: List[List[Any]] = []
    truncated = False
    total = 0
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        for i, row in enumerate(reader):
            if i == 0:
                columns = list(row)
                continue
            total += 1
            if len(rows) >= max_rows:
                truncated = True
                continue
            rows.append(row)
    return {
        "sheet_name": None,
        "sheet_names": [],
        "columns": columns,
        "rows": rows,
        "row_count": total,
        "truncated": truncated,
    }


def read_spreadsheet(
    attachment_id: str,
    sheet_name: Optional[str] = None,
    max_rows: int = 1000,
    user_id: Optional[str] = None,
    **_ignored: Any,
) -> Dict[str, Any]:
    """Read a spreadsheet attachment (XLSX/XLS/ODS/TSV/CSV) and return rows."""
    att, path, err = resolve_attachment(attachment_id, user_id)
    if err is not None:
        return err
    base = {"filename": att.filename}
    try:
        if att.extension == "xlsx":
            base.update(_read_xlsx(path, sheet_name, max_rows))
        elif att.extension == "xls":
            base.update(_read_xls(path, sheet_name, max_rows))
        elif att.extension == "ods":
            base.update(_read_ods(path, sheet_name, max_rows))
        elif att.extension == "tsv":
            base.update(_read_delimited(path, "\t", max_rows))
        elif att.extension == "csv":
            base.update(_read_delimited(path, ",", max_rows))
        else:
            return {"error": {
                "code": "unsupported",
                "message": f"read_spreadsheet does not support .{att.extension}",
            }}
    except Exception as exc:
        logger.exception("spreadsheet parse failed")
        return {"error": {"code": "parse_failed", "message": str(exc)}}
    return base


__all__ = ["read_spreadsheet"]
